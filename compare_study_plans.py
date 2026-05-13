#!/usr/bin/env python3
"""
Compare two EPFL study plan Excel workbooks and highlight the differences.

The script reads the "old" workbook, finds the matching sheets in the "new"
workbook, and annotates the old workbook in-place by highlighting every cell
that changed. A summary sheet called "Diff Summary" is also added to the output
workbook so that additions/removals are easy to review.

Example:
    python compare_study_plans.py \
        --old "1 - Excel SAC 2022/SC-Plan_d_etudes-2022-2023.xlsx" \
        --new "3 - Excel SAC 2023/SC-Plan_d_etudes-2023-2024.xlsx" \
        --output "SC-Plan_d_etudes-2022-2023_with_diffs.xlsx"
"""

from __future__ import annotations

import argparse
import difflib
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

# --- Configuration -----------------------------------------------------------------

# All the header labels we want to recognize. Values are lists of keywords that must
# appear (after accent stripping/lower-casing) for the column to be identified.
FIELD_PATTERNS: Dict[str, Sequence[str]] = {
    "code": ("code", "codes"),
    "course": ("matiere", "matieres", "subjects", "course"),
    # Make the branch type detection strict so we do not mis-detect "Type examen".
    "type": ("type de branche", "type de branches", "branches"),
    "teacher": ("enseignant", "enseignants", "professeur", "teacher", "teachers"),
    "sections": ("section", "sections"),
    "credits": ("credit", "credits", "ects"),
    "semesters": ("semestre", "semestres"),
    "period": ("periode", "period"),
    "language": ("langue",),
    "specializations": (
        "specialisation",
        "specialisations",
        "specialization",
        "specializations",
        "orientation",
        "orientations",
        "mineur",
        "mineurs",
        "minor",
        "minors",
    ),
    "coeff": ("coeff", "coefficient"),
}

# Determines how many columns should be captured for each field. Fields with
# mode == "gap" expand until the next identified header but no further than
# the provided max width.
FIELD_SPAN_RULES: Dict[str, Dict[str, int]] = {
    "period": {"mode": "fixed", "width": 2},
    "type": {"mode": "gap", "max": 12},
    "specializations": {"mode": "gap", "max": 12},
    "semesters": {"mode": "gap", "max": 8},
}

# Colors used for highlights.
MODIFIED_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
REMOVED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
ADDED_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

# Fields intentionally excluded from value comparison.
EXCLUDED_COMPARE_FIELDS = {"teacher", "period", "language", "semesters"}


# --- Helper data structures ---------------------------------------------------------

@dataclass
class SheetLayout:
    """Column ranges (start/end) for each field inside a sheet section."""

    field_ranges: Dict[str, Tuple[int, int]]
    header_labels: Dict[str, str]


@dataclass
class CourseRow:
    """Normalized view of a course row inside a worksheet."""

    sheet: str
    row_idx: int
    display_code: str
    code_key: str
    base_code_key: str
    layout_ranges: Dict[str, Tuple[int, int]]
    field_values: Dict[str, List[Optional[object]]]
    header_labels: Dict[str, str]
    group_key: str
    group_label: str
    group_order: int
    compare_fields: Sequence[str]
    used: bool = field(default=False)


@dataclass
class GroupHeaderRow:
    """A non-course bloc/group row that can still carry comparable values."""

    sheet: str
    row_idx: int
    group_key: str
    group_label: str
    layout_ranges: Dict[str, Tuple[int, int]]
    field_values: Dict[str, List[Optional[object]]]
    compare_fields: Sequence[str]


@dataclass
class FooterRow:
    """A non-empty informational row below the last table on a sheet."""

    sheet: str
    row_idx: int
    values: List[Optional[object]]
    text_key: str


# --- Text utilities -----------------------------------------------------------------

def strip_accents(value: str) -> str:
    """Remove accents/diacritics from a string."""

    normalized = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def normalize_text(value: Optional[object]) -> str:
    """Normalize a header cell so that pattern matching becomes easier."""

    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    value = value.replace("\n", " ")
    value = strip_accents(value).lower()
    return re.sub(r"\s+", " ", value).strip()


def format_display(value: Optional[object]) -> str:
    """Create a nice human-readable representation for summary/comment text."""

    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def normalize_value(value: Optional[object]) -> str:
    """Normalize cell values before comparison."""

    text = format_display(value)
    return text.strip()


def normalize_group_label(text: str) -> str:
    """Normalize group/bloc labels so that similar wording still matches."""

    if not text:
        return ""
    normalized = normalize_text(text)
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return normalized.strip()


def group_kind(group_key: str) -> str:
    """Return the high-level type of a grouping label."""

    if group_key.startswith("bloc"):
        return "bloc"
    if group_key.startswith("block"):
        return "bloc"
    if group_key.startswith("groupe"):
        return "groupe"
    if group_key.startswith("group"):
        return "groupe"
    if group_key.startswith("specialisation"):
        return "specialisation"
    if group_key.startswith("specialization"):
        return "specialisation"
    if group_key.startswith("option"):
        return "option"
    if group_key.startswith("options"):
        return "option"
    if group_key.startswith("gap-"):
        return "gap"
    return "other"


def looks_like_group_header(text: str) -> bool:
    """True only for actual bloc/group/specialization title rows."""

    normalized = normalize_text(text)
    return bool(
        re.match(
            r"^(bloc|block|groupe|group|specialisation|specialization|option|options)\b",
            normalized,
        )
    )


def looks_like_boundary_header(text: str) -> bool:
    """Rows that should split logical groups even if they are not bloc/group titles."""

    normalized = normalize_text(text)
    return bool(
        re.match(
            r"^(optional courses|cours optionnels|total credits|total des credits|total des crédits|notes?\b|remarques?\b)",
            normalized,
        )
    )


CODE_KEY_PATTERN = re.compile(r"[^0-9A-Za-z]+")


def canonical_code(value: Optional[object]) -> Tuple[str, str, str]:
    """Return original display string plus normalized keys for matching."""

    if value is None:
        return "", "", ""
    text = format_display(value)
    text = text.replace("–", "-").replace("—", "-")
    text = text.replace("\xa0", " ").strip()
    base = text.split("(")[0].strip()
    key = CODE_KEY_PATTERN.sub("", text).upper()
    base_key = CODE_KEY_PATTERN.sub("", base).upper()
    return text, key, base_key


def looks_like_course_code(code: str) -> bool:
    """Heuristic to decide if a given string is a course code."""

    if not code:
        return False
    lowered = code.lower().strip()
    # Course codes never contain spaces; this filters out note/comment rows.
    if re.search(r"\s", lowered):
        return False
    if lowered.startswith("total") or lowered.startswith("totaux"):
        return False
    if lowered.startswith("bloc"):
        return False
    return bool(re.search(r"\d", code))


# --- Header/layout detection --------------------------------------------------------

def detect_header(
    row_values: Sequence[Optional[object]],
) -> Optional[Tuple[Dict[str, int], Dict[str, str]]]:
    """Return a mapping from field name to column index plus header labels."""

    mapping: Dict[str, int] = {}
    labels: Dict[str, str] = {}
    for idx, value in enumerate(row_values, start=1):
        normalized = normalize_text(value)
        if not normalized:
            continue
        for field, keywords in FIELD_PATTERNS.items():
            if field in mapping:
                continue
            if field == "code":
                # Avoid false positives like "codesign" in course names while
                # accepting both "Code" and "Codes" headers.
                matched = bool(re.search(r"\bcodes?\b", normalized))
            elif field == "credits":
                # Prevent accidental match of "SUBJECTS" (contains "...ects").
                matched = bool(re.search(r"\bcredits?\b|\bects\b", normalized))
            else:
                matched = any(keyword in normalized for keyword in keywords)
            if matched:
                mapping[field] = idx
                labels[field] = normalized
    # A real header row should identify code plus at least one other known field.
    if "code" in mapping and len(mapping) >= 2:
        return mapping, labels
    return None


def merged_header_span(ws: Worksheet, header_row: int, start_col: int) -> Optional[Tuple[int, int]]:
    """Return the merged-column span for a header cell if it is part of a horizontal merge."""

    for merged in ws.merged_cells.ranges:
        if (
            merged.min_row <= header_row <= merged.max_row
            and merged.min_col <= start_col <= merged.max_col
            and merged.min_col != merged.max_col
        ):
            return merged.min_col, merged.max_col
    return None


def compute_field_ranges(
    ws: Worksheet, header_row: int, header_positions: Dict[str, int], max_column: int
) -> Dict[str, Tuple[int, int]]:
    """Expand header positions into column ranges."""

    sorted_fields = sorted(header_positions.items(), key=lambda item: item[1])
    ranges: Dict[str, Tuple[int, int]] = {}
    for index, (field, start_col) in enumerate(sorted_fields):
        next_start = (
            sorted_fields[index + 1][1] if index + 1 < len(sorted_fields) else max_column + 1
        )
        merged_span = merged_header_span(ws, header_row, start_col)
        if merged_span is not None:
            _, end_col = merged_span
        else:
            rule = FIELD_SPAN_RULES.get(field, {"mode": "fixed", "width": 1})
            if rule["mode"] == "fixed":
                end_col = start_col + rule.get("width", 1) - 1
            else:  # "gap"
                max_width = rule.get("max", 1)
                gap_end = next_start - 1 if next_start > start_col else max_column
                end_col = min(gap_end, start_col + max_width - 1)
        end_col = min(end_col, max_column)
        if end_col < start_col:
            end_col = start_col
        ranges[field] = (start_col, end_col)
    return ranges


def slice_row_values(
    row_values: Sequence[Optional[object]], start: int, end: int
) -> List[Optional[object]]:
    """Return the values from start..end (1-based coordinates)."""

    result: List[Optional[object]] = []
    for col in range(start, end + 1):
        idx = col - 1
        result.append(row_values[idx] if idx < len(row_values) else None)
    return result


def parse_group_headers(ws: Worksheet) -> List[GroupHeaderRow]:
    """Extract bloc/group rows that also carry comparable values such as credits."""

    headers: List[GroupHeaderRow] = []
    current_layout: Optional[SheetLayout] = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = list(row)
        header_detection = detect_header(row_values)
        if header_detection:
            header_map, header_labels = header_detection
            field_ranges = compute_field_ranges(ws, row_idx, header_map, ws.max_column)
            current_layout = SheetLayout(field_ranges=field_ranges, header_labels=header_labels)
            continue
        if not current_layout or "code" not in current_layout.field_ranges:
            continue
        code_start, code_end = current_layout.field_ranges["code"]
        code_cell = slice_row_values(row_values, code_start, code_end)[0]
        course_cell = None
        if "course" in current_layout.field_ranges:
            c_start, c_end = current_layout.field_ranges["course"]
            course_cell = slice_row_values(row_values, c_start, c_end)[0]
        course_text = normalize_value(course_cell)
        code_text = normalize_value(code_cell)
        marker_text = course_text or code_text
        if not marker_text or looks_like_course_code(marker_text):
            continue
        if not (looks_like_group_header(marker_text) or looks_like_boundary_header(marker_text)):
            continue
        field_values = {
            field: slice_row_values(row_values, start, end)
            for field, (start, end) in current_layout.field_ranges.items()
            if field != "code"
        }
        compare_fields = []
        for field, values in field_values.items():
            if field in EXCLUDED_COMPARE_FIELDS:
                continue
            if any(normalize_value(v) for v in values):
                compare_fields.append(field)
        headers.append(
            GroupHeaderRow(
                sheet=ws.title,
                row_idx=row_idx,
                group_key=normalize_group_label(marker_text),
                group_label=marker_text,
                layout_ranges=current_layout.field_ranges,
                field_values=field_values,
                compare_fields=compare_fields,
            )
        )
    return headers


def footer_row_key(values: Sequence[Optional[object]]) -> str:
    """Normalize a whole informational row for matching."""

    parts = [normalize_text(v) for v in values if normalize_text(v)]
    return " | ".join(parts)


def extract_footer_rows(ws: Worksheet) -> List[FooterRow]:
    """Extract non-empty rows below the last structured table content."""

    courses = parse_sheet(ws)
    group_headers = parse_group_headers(ws)
    last_structured_row = 0
    if courses:
        last_structured_row = max(last_structured_row, max(r.row_idx for r in courses))
    if group_headers:
        last_structured_row = max(last_structured_row, max(r.row_idx for r in group_headers))

    rows: List[FooterRow] = []
    for row_idx in range(last_structured_row + 1, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if not any(normalize_value(v) for v in values):
            continue
        rows.append(
            FooterRow(
                sheet=ws.title,
                row_idx=row_idx,
                values=values,
                text_key=footer_row_key(values),
            )
        )
    return rows


def parse_sheet(ws: Worksheet) -> List[CourseRow]:
    """Extract all course rows from a worksheet."""

    courses: List[CourseRow] = []
    current_layout: Optional[SheetLayout] = None
    current_group_key: str = ""
    current_group_label: str = ""
    group_counter = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = list(row)
        header_detection = detect_header(row_values)
        if header_detection:
            header_map, header_labels = header_detection
            field_ranges = compute_field_ranges(ws, row_idx, header_map, ws.max_column)
            current_layout = SheetLayout(field_ranges=field_ranges, header_labels=header_labels)
            current_group_key = ""
            current_group_label = ""
            continue
        if not current_layout or "code" not in current_layout.field_ranges:
            continue
        code_start, code_end = current_layout.field_ranges["code"]
        code_cell = slice_row_values(row_values, code_start, code_end)[0]
        course_cell = None
        if "course" in current_layout.field_ranges:
            c_start, c_end = current_layout.field_ranges["course"]
            course_cell = slice_row_values(row_values, c_start, c_end)[0]
        display_code, code_key, base_code = canonical_code(code_cell)
        # Update grouping markers (Bloc/Groupe/Spécialisation/etc.).
        # Some workbooks place group labels in "Matières", others in "Code".
        course_text = normalize_value(course_cell)
        code_text = normalize_value(code_cell)
        marker_text = course_text or code_text
        normalized_marker_text = normalize_text(marker_text)
        if marker_text and not looks_like_course_code(marker_text) and (
            looks_like_group_header(marker_text) or looks_like_boundary_header(marker_text)
        ):
            group_counter += 1
            current_group_key = normalize_group_label(marker_text)
            current_group_label = marker_text
            continue
        # Treat a fully blank separator row as a new anonymous group boundary.
        if not display_code and not course_text:
            group_counter += 1
            current_group_key = f"gap-{group_counter}"
            current_group_label = ""
            continue
        if not code_key or not looks_like_course_code(display_code):
            continue
        # Capture all fields present in the detected header (besides code); keep them
        # for comparison. Because the user aligned the tables, we can compare every
        # non-code column we detect.
        fields_to_capture = set(current_layout.field_ranges.keys())
        field_ranges = {
            field: rng
            for field, rng in current_layout.field_ranges.items()
            if field in fields_to_capture
        }
        field_values = {
            field: slice_row_values(row_values, start, end)
            for field, (start, end) in field_ranges.items()
        }
        compare_fields = [
            f
            for f in field_ranges.keys()
            if f != "code" and f not in EXCLUDED_COMPARE_FIELDS
        ]
        courses.append(
            CourseRow(
                sheet=ws.title,
                row_idx=row_idx,
                display_code=display_code,
                code_key=code_key,
                base_code_key=base_code,
                layout_ranges=field_ranges,
                field_values=field_values,
                header_labels=current_layout.header_labels,
                group_key=current_group_key,
                group_label=current_group_label,
                group_order=group_counter,
                compare_fields=compare_fields,
            )
        )
    return courses


# --- Sheet comparison ----------------------------------------------------------------

def select_best_match(old_row: CourseRow, candidates: List[CourseRow]) -> int:
    """Pick the most similar candidate row."""

    best_idx = 0
    best_score = -1
    for idx, candidate in enumerate(candidates):
        if candidate.used:
            continue
        score = 0
        fields = [f for f in old_row.compare_fields if f in candidate.compare_fields]
        for field in fields:
            old_vals = old_row.field_values.get(field)
            new_vals = candidate.field_values.get(field)
            if not old_vals or not new_vals:
                continue
            normalized_old = [normalize_value(val) for val in old_vals]
            normalized_new = [normalize_value(val) for val in new_vals]
            score += sum(1 for a, b in zip(normalized_old, normalized_new) if a == b)
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def find_name_match(old_row: CourseRow, candidates: List[CourseRow], threshold: float = 0.7) -> Optional[CourseRow]:
    """Fallback: match by course name similarity when codes don't align."""

    def course_name(row: CourseRow) -> str:
        vals = row.field_values.get("course", [])
        return normalize_text(vals[0]) if vals else ""

    old_name = course_name(old_row)
    best = None
    best_score = threshold
    for cand in candidates:
        score = difflib.SequenceMatcher(None, old_name, course_name(cand)).ratio()
        if score > best_score:
            best_score = score
            best = cand
    return best


def annotate_cell(cell, fill: PatternFill, message: str) -> None:
    """Apply highlighting/comment to a cell."""

    # If this is part of a merged cell range, apply formatting/comment to the anchor.
    if isinstance(cell, MergedCell):
        ws = cell.parent
        anchor = None
        for merged in ws.merged_cells.ranges:
            if (
                merged.min_row <= cell.row <= merged.max_row
                and merged.min_col <= cell.column <= merged.max_col
            ):
                anchor = ws.cell(row=merged.min_row, column=merged.min_col)
                break
        if anchor is None:
            return
        cell = anchor
    cell.fill = fill
    if message:
        if cell.comment:
            text = f"{cell.comment.text}\n{message}"
        else:
            text = message
        cell.comment = Comment(text, "Diff")


def insert_new_row(
    target_ws: Worksheet,
    template_row: CourseRow,
    target_layout_ranges: Dict[str, Tuple[int, int]],
    diff_log: List[Dict[str, str]],
    note: str,
    insert_at: Optional[int] = None,
) -> None:
    """Insert a new row into the target worksheet using data from template_row."""

    insert_at = insert_at or min(template_row.row_idx, target_ws.max_row + 1)
    target_ws.insert_rows(insert_at)
    for field, (start, end) in target_layout_ranges.items():
        values = template_row.field_values.get(field, [])
        width = end - start + 1
        for offset in range(width):
            value = values[offset] if offset < len(values) else None
            col = start + offset
            cell = target_ws.cell(row=insert_at, column=col)
            cell.value = value
            annotate_cell(cell, ADDED_FILL, note)
    template_row.used = True
    diff_log.append(
        {
            "sheet": template_row.sheet,
            "code": template_row.display_code,
            "field": "row",
            "old": "",
            "new": "",
            "details": note,
        }
    )


def compare_rows(
    ws: Worksheet,
    old_row: CourseRow,
    new_row: CourseRow,
    diff_log: List[Dict[str, str]],
    row_offset: int,
) -> int:
    """Highlight per-field differences between two course rows."""

    fields = [f for f in old_row.compare_fields if f in new_row.compare_fields]
    for field in fields:
        if field not in old_row.field_values or field not in new_row.field_values:
            continue
        old_values = old_row.field_values[field]
        new_values = new_row.field_values[field]
        start_col, _ = old_row.layout_ranges[field]
        max_len = max(len(old_values), len(new_values))
        for offset in range(max_len):
            old_val = old_values[offset] if offset < len(old_values) else None
            new_val = new_values[offset] if offset < len(new_values) else None
            if normalize_value(old_val) == normalize_value(new_val):
                continue
            column = start_col + offset
            cell = ws.cell(row=old_row.row_idx + row_offset, column=column)
            annotation = f"value: {format_display(new_val)}"
            annotate_cell(cell, MODIFIED_FILL, annotation)
            diff_log.append(
                {
                    "sheet": old_row.sheet,
                    "code": old_row.display_code,
                    "field": field,
                    "old": format_display(old_val),
                    "new": format_display(new_val),
                    "details": "value changed",
                }
            )
    return 0


def compare_group_header_rows(
    ws: Worksheet,
    old_header: GroupHeaderRow,
    new_header: GroupHeaderRow,
    diff_log: List[Dict[str, str]],
    row_offset: int,
) -> None:
    """Compare values present on a bloc/group title row, such as total credits."""

    fields = [f for f in old_header.compare_fields if f in new_header.compare_fields]
    for field in fields:
        old_values = old_header.field_values.get(field, [])
        new_values = new_header.field_values.get(field, [])
        if not old_values and not new_values:
            continue
        start_col, _ = old_header.layout_ranges[field]
        max_len = max(len(old_values), len(new_values))
        for offset in range(max_len):
            old_val = old_values[offset] if offset < len(old_values) else None
            new_val = new_values[offset] if offset < len(new_values) else None
            if normalize_value(old_val) == normalize_value(new_val):
                continue
            column = start_col + offset
            cell = ws.cell(row=old_header.row_idx + row_offset, column=column)
            annotate_cell(cell, MODIFIED_FILL, f"value: {format_display(new_val)}")
            diff_log.append(
                {
                    "sheet": old_header.sheet,
                    "code": old_header.group_label,
                    "field": field,
                    "old": format_display(old_val),
                    "new": format_display(new_val),
                    "details": "group/header value changed",
                }
            )


def mark_removed_row(
    ws: Worksheet, row: CourseRow, diff_log: List[Dict[str, str]], row_offset: int
) -> None:
    """Highlight an entire course row that disappeared in the new workbook."""

    for field, (start, end) in row.layout_ranges.items():
        for col in range(start, end + 1):
            cell = ws.cell(row=row.row_idx + row_offset, column=col)
            annotate_cell(cell, REMOVED_FILL, "No longer present")
    diff_log.append(
        {
            "sheet": row.sheet,
            "code": row.display_code,
            "field": "row",
            "old": "",
            "new": "",
            "details": "course removed",
        }
    )


def annotate_footer_row(
    ws: Worksheet,
    row_idx: int,
    values: Sequence[Optional[object]],
    fill: PatternFill,
    message: str,
) -> None:
    """Highlight all non-empty cells of a footer/info row."""

    for col_idx, value in enumerate(values, start=1):
        if not normalize_value(value):
            continue
        annotate_cell(ws.cell(row=row_idx, column=col_idx), fill, message)


def compare_footer_row_values(
    ws: Worksheet,
    old_row: FooterRow,
    new_row: FooterRow,
    diff_log: List[Dict[str, str]],
) -> None:
    """Compare two informational footer rows cell-by-cell."""

    max_len = max(len(old_row.values), len(new_row.values))
    changed = False
    for offset in range(max_len):
        old_val = old_row.values[offset] if offset < len(old_row.values) else None
        new_val = new_row.values[offset] if offset < len(new_row.values) else None
        if normalize_value(old_val) == normalize_value(new_val):
            continue
        changed = True
        cell = ws.cell(row=old_row.row_idx, column=offset + 1)
        annotate_cell(cell, MODIFIED_FILL, f"value: {format_display(new_val)}")
    if changed:
        diff_log.append(
            {
                "sheet": old_row.sheet,
                "code": "footer",
                "field": "note",
                "old": old_row.text_key,
                "new": new_row.text_key,
                "details": "footer/note row changed",
            }
        )


def compare_footer_rows(old_ws: Worksheet, new_ws: Worksheet, diff_log: List[Dict[str, str]]) -> None:
    """Compare text rows below the last table and add/remove/update as needed."""

    old_footer = extract_footer_rows(old_ws)
    new_footer = extract_footer_rows(new_ws)
    old_keys = [r.text_key for r in old_footer]
    new_keys = [r.text_key for r in new_footer]
    matcher = difflib.SequenceMatcher(None, old_keys, new_keys)
    inserted_offset = 0

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        if tag == "replace":
            old_slice = old_footer[i1:i2]
            new_slice = new_footer[j1:j2]
            shared = min(len(old_slice), len(new_slice))
            for idx in range(shared):
                compare_footer_row_values(old_ws, old_slice[idx], new_slice[idx], diff_log)
            for row in old_slice[shared:]:
                target_row = row.row_idx + inserted_offset
                annotate_footer_row(old_ws, target_row, row.values, REMOVED_FILL, "No longer present")
                diff_log.append(
                    {
                        "sheet": row.sheet,
                        "code": "footer",
                        "field": "note",
                        "old": row.text_key,
                        "new": "",
                        "details": "footer/note row removed",
                    }
                )
            insert_pos = (
                old_slice[-1].row_idx + inserted_offset + 1
                if old_slice
                else (old_footer[i1 - 1].row_idx + inserted_offset + 1 if i1 > 0 else old_ws.max_row + 1)
            )
            for row in new_slice[shared:]:
                old_ws.insert_rows(insert_pos)
                for col_idx, value in enumerate(row.values, start=1):
                    if value is None:
                        continue
                    cell = old_ws.cell(row=insert_pos, column=col_idx)
                    cell.value = value
                    annotate_cell(cell, ADDED_FILL, "New footer/note row")
                diff_log.append(
                    {
                        "sheet": row.sheet,
                        "code": "footer",
                        "field": "note",
                        "old": "",
                        "new": row.text_key,
                        "details": "footer/note row added",
                    }
                )
                insert_pos += 1
                inserted_offset += 1
        elif tag == "delete":
            for row in old_footer[i1:i2]:
                target_row = row.row_idx + inserted_offset
                annotate_footer_row(old_ws, target_row, row.values, REMOVED_FILL, "No longer present")
                diff_log.append(
                    {
                        "sheet": row.sheet,
                        "code": "footer",
                        "field": "note",
                        "old": row.text_key,
                        "new": "",
                        "details": "footer/note row removed",
                    }
                )
        elif tag == "insert":
            insert_pos = old_ws.max_row + 1
            if i1 < len(old_footer):
                insert_pos = old_footer[i1].row_idx + inserted_offset
            for row in new_footer[j1:j2]:
                old_ws.insert_rows(insert_pos)
                for col_idx, value in enumerate(row.values, start=1):
                    if value is None:
                        continue
                    cell = old_ws.cell(row=insert_pos, column=col_idx)
                    cell.value = value
                    annotate_cell(cell, ADDED_FILL, "New footer/note row")
                diff_log.append(
                    {
                        "sheet": row.sheet,
                        "code": "footer",
                        "field": "note",
                        "old": "",
                        "new": row.text_key,
                        "details": "footer/note row added",
                    }
                )
                insert_pos += 1
                inserted_offset += 1


def compare_sheet(
    old_ws: Worksheet,
    new_ws: Worksheet,
    old_rows: List[CourseRow],
    new_rows: List[CourseRow],
    diff_log: List[Dict[str, str]],
) -> None:
    """Compare every course row in a sheet."""

    if not old_rows:
        return
    old_headers = parse_group_headers(old_ws)
    new_headers = parse_group_headers(new_ws)
    global_code_index = defaultdict(list)
    global_base_index = defaultdict(list)
    for row in new_rows:
        global_code_index[row.code_key].append(row)
        global_base_index[row.base_code_key].append(row)
    # Group rows by bloc/groupe to ensure adds stay within their block.
    def group_rows(rows: List[CourseRow]):
        grouped = defaultdict(list)
        order = []
        for r in sorted(rows, key=lambda r: (r.group_order, r.row_idx)):
            key = r.group_key or ""
            grouped[key].append(r)
            if key not in order:
                order.append(key)
        return grouped, order

    grouped_old, order_old = group_rows(old_rows)
    grouped_new, order_new = group_rows(new_rows)
    default_old_layout_ranges = old_rows[0].layout_ranges
    old_header_map = {h.group_key: h for h in old_headers}
    new_header_map = {h.group_key: h for h in new_headers}

    def map_group_keys(old_keys: List[str], new_keys: List[str]) -> Tuple[Dict[str, Optional[str]], List[str]]:
        remaining = list(new_keys)
        mapping: Dict[str, Optional[str]] = {}
        for ok in old_keys:
            if ok in remaining:
                mapping[ok] = ok
                remaining = [k for k in remaining if k != ok]
                continue
            # Synthetic gap groups should not be fuzzy-matched by label.
            if group_kind(ok) == "gap":
                mapping[ok] = None
                continue
            best_key = None
            best_score = 0.0
            ok_kind = group_kind(ok)
            for nk in list(remaining):
                nk_kind = group_kind(nk)
                # Keep matching within the same grouping family.
                if ok_kind != "gap" and nk_kind != "gap" and ok_kind != nk_kind:
                    continue
                score = difflib.SequenceMatcher(None, ok, nk).ratio()
                if score > best_score:
                    best_score = score
                    best_key = nk
            if best_key and best_score >= 0.6:
                mapping[ok] = best_key
                remaining = [k for k in remaining if k != best_key]
            else:
                mapping[ok] = None
        # If labels changed a lot, map remaining bloc/group sections by course-code overlap.
        for ok in old_keys:
            if mapping.get(ok) is not None:
                continue
            ok_kind = group_kind(ok)
            old_codes = {r.base_code_key for r in grouped_old.get(ok, []) if r.base_code_key}
            if not old_codes:
                continue
            best_key = None
            best_overlap = 0
            for nk in list(remaining):
                nk_kind = group_kind(nk)
                if ok_kind != "gap" and nk_kind != "gap" and ok_kind != nk_kind:
                    continue
                new_codes = {r.base_code_key for r in grouped_new.get(nk, []) if r.base_code_key}
                overlap = len(old_codes & new_codes)
                if overlap > best_overlap:
                    best_overlap = overlap
                    best_key = nk
            if best_key and best_overlap > 0:
                mapping[ok] = best_key
                remaining = [k for k in remaining if k != best_key]
        # Fallback: map any still-unmatched old groups to remaining new groups by order.
        remaining_order = list(remaining)
        for ok in old_keys:
            # Only force positional fallback for synthetic gap groups.
            if mapping.get(ok) is None and group_kind(ok) == "gap" and remaining_order:
                chosen = remaining_order.pop(0)
                mapping[ok] = chosen
                remaining = [k for k in remaining if k != chosen]
        return mapping, remaining

    group_mapping, unmatched_new_groups = map_group_keys(order_old, order_new)

    cumulative_offset = 0

    for group_key in order_old:
        old_group_rows = grouped_old.get(group_key, [])
        mapped_new_key = group_mapping.get(group_key)
        new_group_rows = grouped_new.get(mapped_new_key, []) if mapped_new_key else []

        # Reset usage per bloc to avoid cross-bloc depletion.
        for r in new_group_rows:
            r.used = False

        code_index = defaultdict(list)
        base_index = defaultdict(list)
        for row in new_group_rows:
            code_index[row.code_key].append(row)
            base_index[row.base_code_key].append(row)

        group_start = min(r.row_idx for r in old_group_rows) if old_group_rows else old_ws.max_row
        group_end = max(r.row_idx for r in old_group_rows) if old_group_rows else group_start
        inserted_here = 0

        old_header = old_header_map.get(group_key)
        new_header = new_header_map.get(mapped_new_key) if mapped_new_key else None
        if old_header and new_header:
            compare_group_header_rows(
                old_ws, old_header, new_header, diff_log, cumulative_offset + inserted_here
            )

        for old_row in old_group_rows:
            current_offset = cumulative_offset + inserted_here
            candidates = [row for row in code_index.get(old_row.code_key, []) if not row.used]
            if not candidates:
                candidates = [
                    row for row in base_index.get(old_row.base_code_key, []) if not row.used
                ]
            if not candidates:
                name_candidate = find_name_match(
                    old_row, [r for r in new_group_rows if not r.used]
                )
                if name_candidate:
                    candidates = [name_candidate]
            if not candidates:
                # No match inside this bloc: consider it removed here.
                mark_removed_row(old_ws, old_row, diff_log, current_offset)
                continue
            best_idx = select_best_match(old_row, candidates)
            match = candidates[best_idx]
            match.used = True
            inserted_delta = compare_rows(old_ws, old_row, match, diff_log, current_offset)
            inserted_here += inserted_delta

        for row in new_group_rows:
            if not row.used:
                insert_pos = group_end + cumulative_offset + inserted_here + 1
                target_layout_ranges = (
                    old_group_rows[0].layout_ranges if old_group_rows else default_old_layout_ranges
                )
                insert_new_row(
                    old_ws,
                    row,
                    target_layout_ranges,
                    diff_log,
                    "New course",
                    insert_at=insert_pos,
                )
                inserted_here += 1

        cumulative_offset += inserted_here

    # Handle entirely new groups (if any) by appending at the end in their order.
    for group_key in unmatched_new_groups:
        new_group_rows = grouped_new.get(group_key, [])
        for row in new_group_rows:
            insert_pos = old_ws.max_row + 1
            insert_new_row(
                old_ws,
                row,
                default_old_layout_ranges,
                diff_log,
                f"New course (new bloc {row.group_label or ''})",
                insert_at=insert_pos,
            )

    # Safety net: if any new rows remain unused, append them at the end.
    for row in new_rows:
        if not row.used:
            insert_new_row(
                old_ws,
                row,
                default_old_layout_ranges,
                diff_log,
                f"New course (unmatched bloc {row.group_label or ''})",
                insert_at=old_ws.max_row + 1,
            )


# --- Workbook level orchestration ---------------------------------------------------

def parse_workbook(wb: Workbook) -> Dict[str, List[CourseRow]]:
    """Parse every worksheet in a workbook."""

    data: Dict[str, List[CourseRow]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        data[name] = parse_sheet(ws)
    return data


def normalize_sheet_name(name: str) -> str:
    """Normalize sheet names for fuzzy matching."""

    text = strip_accents(name).lower()
    return re.sub(r"[^a-z0-9]", "", text)


def match_sheets(old_names: Sequence[str], new_names: Sequence[str]) -> Dict[str, Optional[str]]:
    """Map each old sheet name to the closest sheet name in the new workbook."""

    matches: Dict[str, Optional[str]] = {}
    available = set(new_names)
    normalized_new = {name: normalize_sheet_name(name) for name in new_names}
    for old in old_names:
        norm_old = normalize_sheet_name(old)
        exact = [name for name, norm in normalized_new.items() if norm == norm_old and name in available]
        if exact:
            chosen = exact[0]
            matches[old] = chosen
            available.remove(chosen)
            continue
        best_name = None
        best_score = 0.0
        for candidate in available:
            score = difflib.SequenceMatcher(
                None, norm_old, normalized_new[candidate]
            ).ratio()
            if score > best_score:
                best_score = score
                best_name = candidate
        matches[old] = best_name if best_score >= 0.5 else None
        if best_name:
            available.remove(best_name)
    return matches


def write_summary_sheet(wb: Workbook, diff_log: List[Dict[str, str]]) -> None:
    """Create (or replace) the summary sheet containing all the differences."""

    if "Diff Summary" in wb.sheetnames:
        del wb["Diff Summary"]
    ws = wb.create_sheet("Diff Summary")
    ws.append(["Sheet", "Code", "Field", "Old Value", "New Value", "Details"])
    for entry in diff_log:
        ws.append(
            [
                entry.get("sheet", ""),
                entry.get("code", ""),
                entry.get("field", ""),
                entry.get("old", ""),
                entry.get("new", ""),
                entry.get("details", ""),
            ]
        )


def run(old_path: str, new_path: str, output_path: str) -> None:
    """Main orchestration logic."""

    old_wb = load_workbook(old_path)
    new_wb = load_workbook(new_path, data_only=True)
    old_data = parse_workbook(old_wb)
    new_data = parse_workbook(new_wb)
    diff_log: List[Dict[str, str]] = []

    sheet_mapping = match_sheets(old_wb.sheetnames, new_wb.sheetnames)
    for old_name, new_name in sheet_mapping.items():
        if new_name is None:
            continue
        compare_sheet(
            old_wb[old_name],
            new_wb[new_name],
            old_data.get(old_name, []),
            new_data.get(new_name, []),
            diff_log,
        )
        compare_footer_rows(old_wb[old_name], new_wb[new_name], diff_log)

    write_summary_sheet(old_wb, diff_log)
    old_wb.save(output_path)


# --- CLI ----------------------------------------------------------------------------

def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Highlight differences between two study plan Excel workbooks.")
    parser.add_argument("--old", required=True, help="Path to the reference workbook")
    parser.add_argument("--new", required=True, help="Path to the updated workbook")
    parser.add_argument(
        "--output",
        required=True,
        help="Path of the Excel file that will receive the highlighted changes.",
    )
    return parser.parse_args(argv)


def main() -> None:
    args = parse_args()
    run(args.old, args.new, args.output)


if __name__ == "__main__":
    main()
